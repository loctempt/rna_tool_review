import os
import openpyxl.writer.excel
import numpy as np
import math
from config import Config
import sys
import matplotlib as mpl
mpl.use('Agg')

from matplotlib import pyplot as plt 

def cal_y(x,numlist,lig_num):
    print('lig_num:{}\treal_lig_num:{}'.format(lig_num, len(list(filter(lambda x: x=='ligand', numlist)))))
    y=np.arange(0,100,0.001)
    for i in range(len(x)):
        sub_arr = numlist[:int(x[i] / 100 * len(numlist))]
        filtered = list(filter(lambda x: x == 'ligand', sub_arr))
        ligand_cnt = len(filtered)
        # for j in sub_arr:
        #     if j=='ligand':
        #         ligand_cnt+=1
        if len(sub_arr)!=0:
            y[i]=(float(ligand_cnt)/lig_num)*100
    return y

# TODO 取消input pr type pr name 由shell导入
inputPath=Config.DIR_PATH
# protein_type = $1
# gpcr kinse
protein_class = sys.argv[1]
# aa2ar 
# protein_type,protein=map(str,input('please input protein type & name: \n').split())
cur_pr_type =inputPath+'/'+protein_class+'/dock/output/'
cur_dock_input=inputPath+'/'+protein_class+'/dock/input/'
        
workbook = openpyxl.Workbook()

fig = plt.figure(num=1, figsize=(15, 8),dpi=80) 
plt.title(protein_class+" Roc Curve") 
plt.xlabel("Decoy Founds%")
plt.ylabel("Ligand Founds%")
lig_cnt=0

# itero = 0
for iter in ['flex','rigid']:   
    print(iter)
    worksheet = workbook.create_sheet(protein_class+'_'+iter,0)
    # res_dict = [] 
    res_dict={}
    for jter in ['ligand','decoy']:
        print(jter)
        molecule_list = os.listdir(cur_pr_type+str(iter)+'/'+str(jter))
        # if jter=='ligand':
        #     lig_cnt=len(molecule_list)
        for molecule in molecule_list:
            # print(molecule)

            # 对input小分子的分子id匹配
            #  将res_record改为dict key=molecule id ()，value= ((jter),1(molecule id),dock_score)
            if not os.path.isdir(cur_pr_type+str(iter)+'/'+str(jter)+'/'+str(molecule)):
                continue
            with open(cur_pr_type+str(iter)+'/'+str(jter)+'/'+str(molecule)+'/log.txt','r') as cur_file,open(cur_dock_input+str(jter)+'/'+str(molecule)+'.pdbqt','r') as input_dock:
                print(iter,' ',jter,' ',str(molecule))
                ligand_id = input_dock.readlines()[0].split()[3]   
                dock_score = min([float(record.split()[1]) for record in cur_file.readlines()[24:-1]])
		#print('小分子 数据库内id: ',ligand_id)
                if ligand_id not in res_dict.keys() or ligand_id in res_dict.keys() and dock_score <= res_dict[ligand_id][2]:
                    res_dict[ligand_id] = (jter,molecule,dock_score)
                    if jter == 'ligand':
                        lig_cnt += 1
                    
 #   print(res_dict)
    res_dict = sorted(res_dict.items(),key=(lambda x: x[1][2]))
   # print(res_dict)
    x = np.arange(0,100,0.001) 
    y = cal_y(x,[rec[1][0] for rec in res_dict],lig_cnt)
    plt.plot(x,y,label=iter) 

    cnt=1
    for res_record in res_dict:
        worksheet.cell(cnt,1).value=str(res_record[0])
        print(str(res_record[0]),str(res_record[1][0]),str(res_record[1][1]),str(res_record[1][2]))
        worksheet.cell(cnt,2).value=str(res_record[1][0])
        worksheet.cell(cnt,3).value=str(res_record[1][1])
        worksheet.cell(cnt,4).value=float(res_record[1][2])
        cnt+=1

workbook.save(cur_pr_type+protein_class+'.xlsx')
plt.legend()
# plt.show()
plt.savefig(cur_pr_type+protein_class+'.svg')



# decoyresultfilePath=r'C:\Users\lenovo\Desktop\biyesheji\akt1\virtual screen\autodockVina\output\rigid\decoy\result'
# ligandresultfilePath=r'C:\Users\lenovo\Desktop\biyesheji\akt1\virtual screen\autodockVina\output\rigid\ligand\result'
# decoyMolsPath=r'C:\Users\lenovo\Desktop\biyesheji\akt1\virtual screen\autodockVina\dockprep\decoy\mol2'
# ligandMolsPath=r'C:\Users\lenovo\Desktop\biyesheji\akt1\virtual screen\autodockVina\dockprep\ligand\mol2'
# outPath=r'C:\Users\lenovo\Desktop\biyesheji\akt1\virtual screen\autodockVina\output\rigid'

# decoyList=os.listdir(decoyresultfilePath)
# ligandList=os.listdir(ligandresultfilePath)
# workbook = openpyxl.Workbook()
# worksheet = workbook.create_sheet('akt1',0)
# tempdict={}

# for decoy in decoyList:
#     decoy=decoy.split('.')
#     with open(decoyresultfilePath+'/'+decoy[0]+'.'+decoy[1],'r') as decoyresFile,open(decoyMolsPath+'/'+decoy[0]+'.mol2','r')as decoymol2File:
#         templist=[]
#         for line in decoyresFile.readlines():
#             line=line.split()
#             templist.append('decoy')
#             templist.append(int(decoy[0]))
#             templist.append(float(line[1]))
            

#         for line in decoymol2File.readlines():
#             if line[0:4]=='ZINC':
#                 if tempdict.__contains__(str(line)):
#                     if tempdict[str(line)][2]>templist[2]:
#                         tempdict[str(line)]=templist
#                 else:
#                     tempdict[str(line)]=templist
#                 break
            

# for ligand in ligandList:
#     ligand=ligand.split('.')
#     with open(ligandresultfilePath+'/'+ligand[0]+'.result','r') as ligandresFile,open(ligandMolsPath+'/'+ligand[0]+'.mol2','r') as ligandmol2File:
#         templist=[]
#         for line in ligandresFile.readlines():
#             line=line.split()
#             templist.append('ligand')
#             templist.append(int(ligand[0]))
#             templist.append(float(line[1]))
            

#         for line in ligandmol2File.readlines():
#             if line[0:6]=='CHEMBL':
#                 if tempdict.__contains__(str(line)) :
#                     if tempdict[str(line)][2]>templist[2]:
#                         tempdict[str(line)]=templist
#                 else:
#                     tempdict[str(line)]=templist
#                 break
            


# cnt=1
# print(len(tempdict))

# for key in tempdict:
#     worksheet.cell(cnt,1).value=key
#     temp=tempdict[key]
#     print(key,cnt,tempdict[key][1])
#     worksheet.cell(cnt,2).value=tempdict[key][0]
#     worksheet.cell(cnt,3).value=tempdict[key][1]
#     worksheet.cell(cnt,4).value=tempdict[key][2]
#     cnt+=1
# workbook.save(outPath+'/akt1_rigid.xlsx')


